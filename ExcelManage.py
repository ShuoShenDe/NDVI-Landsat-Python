from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
import tkinter as tk
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg, NavigationToolbar2Tk)

#操纵Excel

#写文件
def writefile(year, value):
    row = int(year) - 2010 + 2      #计算结果在第几行
    dest_filename = 'NDVI.xlsx'      #文件名
    wb = load_workbook(filename=dest_filename)  #打开文件
    sheet_ranges = wb['Sheet1']   #打开Sheet1
    print('A' + row.__str__())
    sheet_ranges['A' + row.__str__()] = year   #第一列写入年份
    sheet_ranges['B' + row.__str__()] = value   #第二列写入NDVI值
    wb.save(filename=dest_filename)    #保存

#显示NDVI折线图
def showNDVI(startyear, endyear):
    print("Plot line chart")
    dest_filename = 'NDVI.xlsx'     #文件名
    wb = load_workbook(filename=dest_filename)   #加载工作区
    sheet_ranges = wb['Sheet1']   #读取Sheet1
    value = []
    year = []
    for i in range(int(startyear), int(endyear)+1):   #遍历数据读取，并存入year和value中
        row = i - 2010 + 2
        year.append(i)
        value.append(sheet_ranges['B' + row.__str__()].value)


    #画折线图
    root = tk.Tk()
    root.wm_title("NDVI")  #折线图名称为NDVI
    fig = Figure(figsize=(5, 4), dpi=100)
    fig.add_subplot(111).plot(year, value, marker='', color='red', linewidth=2.4, alpha=0.9)  #传入数据
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    toolbar = NavigationToolbar2Tk(canvas, root)
    toolbar.update()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

    tk.mainloop()


    return year, value
